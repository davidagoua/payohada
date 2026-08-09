"""
Router Import/Export Excel — Variables de Paie (Format QPXL1501)

Ce module permet :
- D'exporter la liste des salariés dans un fichier Excel pré-rempli (format QPXL1501)
  pour faciliter la saisie des variables de paie (heures sup, absences, primes, acomptes).
- D'importer ce fichier rempli pour créer/mettre à jour les variables de paie
  en base de données, prêtes pour le calcul des bulletins.
"""

from __future__ import annotations

import io
import math
from datetime import date, datetime, timezone
from typing import Optional, List

import xlrd
import openpyxl
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side
)
from openpyxl.utils import get_column_letter

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session, joinedload

from app.database import get_db
from app.models.models import (
    Absence,
    Contrat,
    Dossier,
    Etablissement,
    HeureSupplementaire,
    Prime,
    Salarie,
    Utilisateur,
)
from app.services.security import get_current_user

router = APIRouter(tags=["Import/Export Excel"])


# ──────────────────────────────────────────────────────────────
#  HELPERS
# ──────────────────────────────────────────────────────────────

def _check_dossier(dossier_id: int, user_id: int, db: Session) -> Dossier:
    dossier = (
        db.query(Dossier)
        .filter(Dossier.id == dossier_id, Dossier.utilisateur_id == user_id)
        .first()
    )
    if not dossier:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Dossier introuvable ou accès refusé.",
        )
    return dossier


def _get_salaries_actifs(dossier_id: int, db: Session):
    salaries = (
        db.query(Salarie)
        .join(Etablissement, Salarie.etablissement_id == Etablissement.id)
        .filter(
            Etablissement.dossier_id == dossier_id,
            Salarie.is_active == True,
        )
        .options(joinedload(Salarie.contrats))
        .order_by(Salarie.id)
        .all()
    )
    return salaries


def _get_contrat_actif(salarie: Salarie) -> Optional[Contrat]:
    for c in salarie.contrats:
        if c.statut == "actif":
            return c
    return salarie.contrats[0] if salarie.contrats else None


# ──────────────────────────────────────────────────────────────
#  STYLES EXCEL
# ──────────────────────────────────────────────────────────────

BLUE_HEADER   = "1F4E79"
LIGHT_BLUE    = "BDD7EE"
ORANGE_HDR    = "C55A11"
LIGHT_ORANGE  = "FCE4D6"
GREEN_HDR     = "375623"
LIGHT_GREEN   = "E2EFDA"
PURPLE_HDR    = "7030A0"
LIGHT_PURPLE  = "EAD1F5"
WHITE         = "FFFFFF"
GREY_ROW      = "F2F2F2"

thin   = Side(border_style="thin",   color="999999")
medium = Side(border_style="medium", color="444444")
THIN_BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def _make_fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _make_header_cell(ws, row: int, col: int, value, color_hex: str, bold: bool = True):
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill = _make_fill(color_hex)
    cell.font = Font(bold=bold, color=WHITE, size=10)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = THIN_BORDER
    return cell


def _make_subheader_cell(ws, row: int, col: int, value, color_hex: str):
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill = _make_fill(color_hex)
    cell.font = Font(bold=False, color="1F1F1F", size=9, italic=True)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = THIN_BORDER
    return cell


def _make_id_cell(ws, row: int, col: int, value):
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill = _make_fill(LIGHT_BLUE)
    cell.font = Font(bold=True, color="1F4E79", size=10)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = THIN_BORDER
    return cell


def _make_name_cell(ws, row: int, col: int, value):
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill = _make_fill(GREY_ROW)
    cell.font = Font(bold=True, color="1F1F1F", size=10)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = THIN_BORDER
    return cell


def _make_data_cell(ws, row: int, col: int, bg: str = WHITE):
    cell = ws.cell(row=row, column=col, value=None)
    cell.fill = _make_fill(bg)
    cell.font = Font(size=10)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = THIN_BORDER
    return cell


def _set_col_width(ws, col: int, width: float):
    ws.column_dimensions[get_column_letter(col)].width = width


# ──────────────────────────────────────────────────────────────
#  GÉNÉRATION FEUILLES
# ──────────────────────────────────────────────────────────────

def _build_saisie_sheet(ws, salaries, mois, annee, dossier_code):
    ws.merge_cells("B2:O2")
    titre = ws.cell(row=2, column=2, value="Saisie des éléments variables de paie")
    titre.font = Font(bold=True, size=14, color=BLUE_HEADER)
    titre.alignment = Alignment(horizontal="center", vertical="center")

    ws.cell(row=4, column=2, value="Dossier :").font = Font(bold=True, size=10)
    ws.cell(row=4, column=3, value=dossier_code)
    ws.cell(row=4, column=4, value="Période :").font = Font(bold=True, size=10)
    if mois and annee:
        ws.cell(row=4, column=5, value=f"{mois:02d}/{annee}")
    elif annee:
        ws.cell(row=4, column=5, value=str(annee))

    headers_codes  = ["", "Numéro", "Nom/Prénom", ".HBA",       ".BAS",        ".HSB",       "TSAL",        ".HCO",   ".HS1",   ".HS2",   ".HS3",   ".RTP",    "BP01",           "BP02",                  "BP06"]
    headers_labels = ["", "",       "",            "H Salaire",  "Mt Salaire",  "H Bonif.",   "Total Sal.",  "H Comp", "H Sup1", "H Sup2", "H Sup3", "RTT pris","PRIME PROJET",   "PRIME EXCEPTIONNELLE",  "AVANCE / PRIME OBJECTIF"]

    for c, (code, label) in enumerate(zip(headers_codes, headers_labels), start=1):
        if code:
            _make_header_cell(ws, 6, c, code, BLUE_HEADER)
            _make_subheader_cell(ws, 7, c, label, LIGHT_BLUE)

    for i, sal in enumerate(salaries):
        row = 8 + i
        bg = WHITE if i % 2 == 0 else GREY_ROW
        _make_id_cell(ws, row, 2, sal.id)
        _make_name_cell(ws, row, 3, f"{sal.nom} {sal.prenom}")
        for c in range(4, 16):
            _make_data_cell(ws, row, c, bg)

    col_widths = [2, 10, 28, 11, 14, 11, 13, 10, 9, 9, 9, 9, 16, 22, 24]
    for c, w in enumerate(col_widths, start=1):
        _set_col_width(ws, c, w)
    ws.row_dimensions[2].height = 28
    ws.row_dimensions[6].height = 22
    ws.row_dimensions[7].height = 18
    ws.freeze_panes = "D8"


def _build_absences_sheet(ws, salaries, dossier_code):
    ws.merge_cells("B2:J2")
    titre = ws.cell(row=2, column=2, value="Saisie des absences et congés")
    titre.font = Font(bold=True, size=14, color=ORANGE_HDR)
    titre.alignment = Alignment(horizontal="center", vertical="center")

    types_abs = [
        "C - Congés Payés", "0 - Maladie", "1 - A.T.", "2 - Maternité",
        "10 - Abs. Congés payés", "100 - Chômage intempéries", "101 - Activité partielle",
        "3 - Injustifiées", "4 - Congés SS", "5 - Diverses",
    ]
    ws.cell(row=3, column=10, value="Types d'absences").font = Font(bold=True, size=9, color=ORANGE_HDR)
    for i, t in enumerate(types_abs):
        ws.cell(row=4 + i, column=10, value=t).font = Font(size=9)

    headers = ["", "Numéro", "Nom/Prénom", "Début", "Fin", "Type", "Nb Jours", "Nb Heures", ""]
    for c, h in enumerate(headers, start=1):
        if h:
            _make_header_cell(ws, 3, c, h, ORANGE_HDR)

    for i, sal in enumerate(salaries):
        row = 4 + i
        bg = WHITE if i % 2 == 0 else LIGHT_ORANGE
        _make_id_cell(ws, row, 2, sal.id)
        _make_name_cell(ws, row, 3, f"{sal.nom} {sal.prenom}")
        for c in range(4, 9):
            _make_data_cell(ws, row, c, bg)

    col_widths = [2, 10, 28, 14, 14, 10, 11, 11, 2, 22]
    for c, w in enumerate(col_widths, start=1):
        _set_col_width(ws, c, w)
    ws.row_dimensions[2].height = 28
    ws.row_dimensions[3].height = 22
    ws.freeze_panes = "D4"


def _build_rtt_sheet(ws, salaries, dossier_code):
    ws.merge_cells("B2:E2")
    titre = ws.cell(row=2, column=2, value="Saisie des journées de RTT")
    titre.font = Font(bold=True, size=14, color=GREEN_HDR)
    titre.alignment = Alignment(horizontal="center", vertical="center")

    for c, h in enumerate(["", "Numéro", "Nom/Prénom", "Date", "Nb Jours"], start=1):
        if h:
            _make_header_cell(ws, 3, c, h, GREEN_HDR)

    for i, sal in enumerate(salaries):
        row = 4 + i
        bg = WHITE if i % 2 == 0 else LIGHT_GREEN
        _make_id_cell(ws, row, 2, sal.id)
        _make_name_cell(ws, row, 3, f"{sal.nom} {sal.prenom}")
        _make_data_cell(ws, row, 4, bg)
        _make_data_cell(ws, row, 5, bg)

    for c, w in enumerate([2, 10, 30, 16, 12], start=1):
        _set_col_width(ws, c, w)
    ws.row_dimensions[2].height = 28
    ws.freeze_panes = "D4"


def _build_acomptes_sheet(ws, salaries, dossier_code):
    # Titre principal sur la ligne 2 (colonnes B à H fusionnées)
    ws.merge_cells("B2:H2")
    titre = ws.cell(row=2, column=2, value="Saisie des acomptes")
    titre.font = Font(bold=True, size=14, color=PURPLE_HDR)
    titre.alignment = Alignment(horizontal="center", vertical="center")

    # Modes de paiement sur les lignes 3-5 (hors zone fusionnée)
    ws.cell(row=3, column=2, value="Modes de paiement :").font = Font(bold=True, size=9)
    for i, m in enumerate(["1 - Chèque", "2 - Virement", "3 - Espèces"]):
        ws.cell(row=3 + i, column=3, value=m).font = Font(size=9)

    # Note CP sur une ligne séparée
    note = ws.cell(row=3, column=5, value="S'il s'agit d'un acompte sur congés payés, renseigner la colonne prévue à cet effet par Oui (ou O, ou X ...)")
    note.font = Font(size=8, italic=True, color="666666")

    headers = ["", "Numéro", "Nom/Prénom", "Libelle", "Date", "Mode paiement", "Aco. sur CP ?", "Montant"]
    for c, h in enumerate(headers, start=1):
        if h:
            _make_header_cell(ws, 6, c, h, PURPLE_HDR)

    for i, sal in enumerate(salaries):
        row = 7 + i
        bg = WHITE if i % 2 == 0 else LIGHT_PURPLE
        _make_id_cell(ws, row, 2, sal.id)
        _make_name_cell(ws, row, 3, f"{sal.nom} {sal.prenom}")
        for c in range(4, 9):
            _make_data_cell(ws, row, c, bg)

    for c, w in enumerate([2, 10, 28, 25, 14, 16, 14, 14], start=1):
        _set_col_width(ws, c, w)
    ws.row_dimensions[2].height = 28
    ws.row_dimensions[6].height = 22
    ws.freeze_panes = "D7"



# ──────────────────────────────────────────────────────────────
#  ENDPOINT EXPORT
# ──────────────────────────────────────────────────────────────

@router.get(
    "/dossiers/{dossier_id}/export-variables-excel",
    summary="Exporter les salariés au format Excel QPXL1501",
)
def export_variables_excel(
    dossier_id: int,
    mois: Optional[int] = Query(None, ge=1, le=12, description="Mois de la période (1-12)"),
    annee: Optional[int] = Query(None, ge=2000, le=2100, description="Année de la période"),
    db: Session = Depends(get_db),
    current_user: Utilisateur = Depends(get_current_user),
):
    """
    Génère un fichier Excel au format QPXL1501 pré-rempli avec l'id et le nom/prénom
    de chaque salarié actif du dossier. 4 feuilles : Saisie, Absences, RTT, Acomptes.
    """
    dossier  = _check_dossier(dossier_id, current_user.id, db)
    salaries = _get_salaries_actifs(dossier_id, db)

    if not salaries:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Aucun salarié actif trouvé dans ce dossier.",
        )

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws_saisie   = wb.create_sheet("Saisie")
    ws_absences = wb.create_sheet("Absences")
    ws_rtt      = wb.create_sheet("RTT")
    ws_acomptes = wb.create_sheet("Acomptes")

    ws_saisie.sheet_properties.tabColor   = BLUE_HEADER
    ws_absences.sheet_properties.tabColor = ORANGE_HDR
    ws_rtt.sheet_properties.tabColor      = GREEN_HDR
    ws_acomptes.sheet_properties.tabColor = PURPLE_HDR

    _build_saisie_sheet(ws_saisie,    salaries, mois, annee, dossier.code)
    _build_absences_sheet(ws_absences, salaries, dossier.code)
    _build_rtt_sheet(ws_rtt,          salaries, dossier.code)
    _build_acomptes_sheet(ws_acomptes, salaries, dossier.code)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    periode = f"_{mois:02d}_{annee}" if mois and annee else (f"_{annee}" if annee else "")
    filename = f"variables_paie_{dossier.code}{periode}.xlsx"

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ──────────────────────────────────────────────────────────────
#  PARSING HELPERS
# ──────────────────────────────────────────────────────────────

def _cell_float(val) -> Optional[float]:
    if val is None or val == "" or (isinstance(val, float) and math.isnan(val)):
        return None
    try:
        f = float(val)
        return f if f != 0.0 else None
    except (TypeError, ValueError):
        return None


def _cell_str(val) -> str:
    if val is None:
        return ""
    return str(val).strip()


def _parse_date_val(val) -> Optional[date]:
    if val is None or val == "":
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, float):
        from datetime import timedelta
        try:
            return date(1899, 12, 30) + timedelta(days=int(val))
        except Exception:
            return None
    if isinstance(val, str):
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
            try:
                return datetime.strptime(val.strip(), fmt).date()
            except ValueError:
                pass
    return None


def _load_workbook_any(file_bytes: bytes, filename: str):
    if filename.lower().endswith(".xls"):
        wb = xlrd.open_workbook(file_contents=file_bytes)
        return ("xls", wb)
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    return ("xlsx", wb)


def _iter_sheet_rows(wb_info, sheet_name: str):
    fmt, wb = wb_info
    if fmt == "xls":
        if sheet_name not in wb.sheet_names():
            return
        sheet = wb.sheet_by_name(sheet_name)
        for i in range(sheet.nrows):
            yield [sheet.cell_value(i, j) for j in range(sheet.ncols)]
    else:
        if sheet_name not in wb.sheetnames:
            return
        for row in wb[sheet_name].iter_rows(values_only=True):
            yield list(row)


ABSENCE_TYPE_MAP = {
    "C":   "CP",
    "0":   "MAL",
    "1":   "AT",
    "2":   "MAT",
    "10":  "ABS_CP",
    "100": "INTEMP",
    "101": "ACT_PART",
    "3":   "INJ",
    "4":   "CSS",
    "5":   "DIV",
    "6":   "CONG_PARENTAL",
    "7":   "RETARD",
    "8":   "ABS_NON_PAY",
    "11":  "ABS_CONGES",
}


# ──────────────────────────────────────────────────────────────
#  SCHÉMA RÉPONSE IMPORT
# ──────────────────────────────────────────────────────────────

class ImportResult(BaseModel):
    salaries_traites: int = 0
    variables_creees: int = 0
    absences_creees: int = 0
    primes_creees: int = 0
    acomptes_crees: int = 0
    heures_sup_creees: int = 0
    erreurs: List[str] = []
    avertissements: List[str] = []


# ──────────────────────────────────────────────────────────────
#  ENDPOINT IMPORT
# ──────────────────────────────────────────────────────────────

@router.post(
    "/dossiers/{dossier_id}/import-variables-excel",
    response_model=ImportResult,
    summary="Importer les variables de paie depuis un fichier Excel QPXL1501",
)
async def import_variables_excel(
    dossier_id: int,
    mois: int = Query(..., ge=1, le=12, description="Mois de la période (1-12)"),
    annee: int = Query(..., ge=2000, le=2100, description="Année de la période"),
    fichier: UploadFile = File(..., description="Fichier .xls ou .xlsx au format QPXL1501"),
    db: Session = Depends(get_db),
    current_user: Utilisateur = Depends(get_current_user),
):
    """
    Importe les variables de paie depuis un fichier Excel au format QPXL1501.

    - Le **premier champ** de chaque ligne est l'**ID du salarié** dans la base.
    - Les variables existantes pour la période (mois/année) sont **remplacées**.
    - Les bulletins ne sont **pas calculés automatiquement** après l'import.
    """
    _check_dossier(dossier_id, current_user.id, db)

    file_bytes = await fichier.read()
    if not file_bytes:
        raise HTTPException(status_code=400, detail="Le fichier est vide.")

    filename = fichier.filename or "upload.xlsx"
    try:
        wb_info = _load_workbook_any(file_bytes, filename)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Impossible de lire le fichier Excel : {e}")

    result        = ImportResult()
    annee_str     = str(annee)
    salaries_traites: set = set()
    _contrat_cache: dict  = {}

    def _get_contrat_by_salarie_id(salarie_id: int) -> Optional[Contrat]:
        sal = (
            db.query(Salarie)
            .join(Etablissement, Salarie.etablissement_id == Etablissement.id)
            .filter(
                Salarie.id == salarie_id,
                Etablissement.dossier_id == dossier_id,
            )
            .first()
        )
        if not sal:
            return None
        return _get_contrat_actif(sal)

    def get_contrat(salarie_id: int) -> Optional[Contrat]:
        if salarie_id not in _contrat_cache:
            _contrat_cache[salarie_id] = _get_contrat_by_salarie_id(salarie_id)
        return _contrat_cache[salarie_id]

    # ─── Feuille SAISIE ───────────────────────────────────────
    # Colonnes (0-indexed): 1=id, 2=nom, 3=.HBA(ignoré), 4=.BAS(ignoré),
    # 5=.HSB(ignoré), 6=TSAL(ignoré), 7=.HCO, 8=.HS1, 9=.HS2,
    # 10=.HS3, 11=.RTP, 12=BP01, 13=BP02, 14=BP06

    data_row_started = False
    for row in _iter_sheet_rows(wb_info, "Saisie"):
        if len(row) < 3:
            continue
        row_str = [_cell_str(c) for c in row]
        if "Numéro" in row_str and "Nom/Prénom" in row_str:
            data_row_started = True
            continue
        if not data_row_started:
            continue
        if "H Salaire" in row_str or "Mt Salaire" in row_str:
            continue  # sous-en-tête

        raw_id = row[1] if len(row) > 1 else None
        if raw_id is None or raw_id == "":
            continue
        try:
            salarie_id = int(float(str(raw_id).strip()))
        except (ValueError, TypeError):
            continue

        contrat = get_contrat(salarie_id)
        if contrat is None:
            result.avertissements.append(
                f"Saisie — Salarié ID {salarie_id} introuvable ou sans contrat actif."
            )
            continue

        salaries_traites.add(salarie_id)

        # Nettoyer les variables existantes pour la période
        db.query(HeureSupplementaire).filter(
            HeureSupplementaire.contrat_id == contrat.id,
            HeureSupplementaire.mois == mois,
            HeureSupplementaire.annee == annee_str,
        ).delete(synchronize_session=False)
        db.query(Prime).filter(
            Prime.contrat_id == contrat.id,
            Prime.mois == mois,
            Prime.annee == annee_str,
            Prime.code.in_(["HCO", "HS25", "HS50", "HS100", "BP01", "BP02", "BP06"]),
        ).delete(synchronize_session=False)
        db.query(Absence).filter(
            Absence.contrat_id == contrat.id,
            Absence.mois == mois,
            Absence.annee == annee_str,
            Absence.code == "RTT",
        ).delete(synchronize_session=False)

        # .HCO
        val = _cell_float(row[7]) if len(row) > 7 else None
        if val is not None:
            db.add(HeureSupplementaire(contrat_id=contrat.id, code="HCO", nombre=val, mois=mois, annee=annee_str))
            result.heures_sup_creees += 1
            result.variables_creees += 1

        # .HS1 → HS25
        val = _cell_float(row[8]) if len(row) > 8 else None
        if val is not None:
            db.add(HeureSupplementaire(contrat_id=contrat.id, code="HS25", nombre=val, mois=mois, annee=annee_str))
            result.heures_sup_creees += 1
            result.variables_creees += 1

        # .HS2 → HS50
        val = _cell_float(row[9]) if len(row) > 9 else None
        if val is not None:
            db.add(HeureSupplementaire(contrat_id=contrat.id, code="HS50", nombre=val, mois=mois, annee=annee_str))
            result.heures_sup_creees += 1
            result.variables_creees += 1

        # .HS3 → HS100
        val = _cell_float(row[10]) if len(row) > 10 else None
        if val is not None:
            db.add(HeureSupplementaire(contrat_id=contrat.id, code="HS100", nombre=val, mois=mois, annee=annee_str))
            result.heures_sup_creees += 1
            result.variables_creees += 1

        # .RTP → Absence RTT
        val = _cell_float(row[11]) if len(row) > 11 else None
        if val is not None:
            db.add(Absence(contrat_id=contrat.id, code="RTT", nbr_jour_by_user=val, mois=mois, annee=annee_str))
            result.absences_creees += 1
            result.variables_creees += 1

        # BP01
        val = _cell_float(row[12]) if len(row) > 12 else None
        if val is not None:
            db.add(Prime(contrat_id=contrat.id, code="BP01", libelle="PRIME PROJET", montant=val, mois=mois, annee=annee_str))
            result.primes_creees += 1
            result.variables_creees += 1

        # BP02
        val = _cell_float(row[13]) if len(row) > 13 else None
        if val is not None:
            db.add(Prime(contrat_id=contrat.id, code="BP02", libelle="PRIME EXCEPTIONNELLE", montant=val, mois=mois, annee=annee_str))
            result.primes_creees += 1
            result.variables_creees += 1

        # BP06
        val = _cell_float(row[14]) if len(row) > 14 else None
        if val is not None:
            db.add(Prime(contrat_id=contrat.id, code="BP06", libelle="AVANCE / PRIME OBJECTIF", montant=val, mois=mois, annee=annee_str))
            result.primes_creees += 1
            result.variables_creees += 1

    # ─── Feuille ABSENCES ─────────────────────────────────────
    # Colonnes: 1=id, 2=nom, 3=Début, 4=Fin, 5=Type, 6=Nb Jours, 7=Nb Heures

    abs_data_started      = False
    _cleaned_abs_contrats: set = set()

    for row in _iter_sheet_rows(wb_info, "Absences"):
        if len(row) < 3:
            continue
        row_str = [_cell_str(c) for c in row]
        if "Numéro" in row_str and "Nom/Prénom" in row_str:
            abs_data_started = True
            continue
        if not abs_data_started:
            continue

        raw_id = row[1] if len(row) > 1 else None
        if raw_id is None or raw_id == "":
            continue
        try:
            salarie_id = int(float(str(raw_id).strip()))
        except (ValueError, TypeError):
            continue

        contrat = get_contrat(salarie_id)
        if contrat is None:
            result.avertissements.append(f"Absences — Salarié ID {salarie_id} introuvable.")
            continue

        salaries_traites.add(salarie_id)

        if contrat.id not in _cleaned_abs_contrats:
            db.query(Absence).filter(
                Absence.contrat_id == contrat.id,
                Absence.mois == mois,
                Absence.annee == annee_str,
                Absence.code != "RTT",
            ).delete(synchronize_session=False)
            _cleaned_abs_contrats.add(contrat.id)

        date_debut = _parse_date_val(row[3]) if len(row) > 3 else None
        date_fin   = _parse_date_val(row[4]) if len(row) > 4 else None
        type_abs   = _cell_str(row[5]) if len(row) > 5 else ""
        nb_jours   = _cell_float(row[6]) if len(row) > 6 else None
        nb_heures  = _cell_float(row[7]) if len(row) > 7 else None

        if not type_abs and nb_jours is None and nb_heures is None:
            continue

        # Normaliser le code type: "0.0" → "0", "4.0" → "4", "C" → "C"
        raw_type = type_abs.strip()
        if raw_type.endswith(".0") and raw_type[:-2].isdigit():
            raw_type = raw_type[:-2]

        code_abs = ABSENCE_TYPE_MAP.get(raw_type, raw_type or "DIV")

        def _to_dt(d: Optional[date]) -> Optional[datetime]:
            if d is None:
                return None
            return datetime.combine(d, datetime.min.time()).replace(tzinfo=timezone.utc)

        db.add(Absence(
            contrat_id=contrat.id,
            code=code_abs,
            date_debut=_to_dt(date_debut),
            date_fin=_to_dt(date_fin),
            nbr_jour_by_user=nb_jours or 0.0,
            nbr_heure_by_user=nb_heures or 0.0,
            mois=mois,
            annee=annee_str,
        ))
        result.absences_creees += 1
        result.variables_creees += 1

    # ─── Feuille ACOMPTES ─────────────────────────────────────
    # Colonnes: 1=id, 2=nom, 3=Libelle, 4=Date, 5=Mode, 6=AcoCP?, 7=Montant

    acompte_started           = False
    _cleaned_acompte_contrats: set = set()

    for row in _iter_sheet_rows(wb_info, "Acomptes"):
        if len(row) < 3:
            continue
        row_str = [_cell_str(c) for c in row]
        if "Numéro" in row_str and "Montant" in row_str:
            acompte_started = True
            continue
        if not acompte_started:
            continue

        raw_id = row[1] if len(row) > 1 else None
        if raw_id is None or raw_id == "":
            continue
        try:
            salarie_id = int(float(str(raw_id).strip()))
        except (ValueError, TypeError):
            continue

        montant = _cell_float(row[7]) if len(row) > 7 else None
        if montant is None:
            continue

        contrat = get_contrat(salarie_id)
        if contrat is None:
            result.avertissements.append(f"Acomptes — Salarié ID {salarie_id} introuvable.")
            continue

        salaries_traites.add(salarie_id)

        if contrat.id not in _cleaned_acompte_contrats:
            db.query(Prime).filter(
                Prime.contrat_id == contrat.id,
                Prime.mois == mois,
                Prime.annee == annee_str,
                Prime.code == "ACOMPTE",
            ).delete(synchronize_session=False)
            _cleaned_acompte_contrats.add(contrat.id)

        libelle = _cell_str(row[3]) if len(row) > 3 else ""
        if not libelle:
            libelle = "Acompte"

        db.add(Prime(
            contrat_id=contrat.id,
            code="ACOMPTE",
            libelle=libelle,
            montant=montant,
            mois=mois,
            annee=annee_str,
        ))
        result.acomptes_crees += 1
        result.primes_creees  += 1
        result.variables_creees += 1

    try:
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Erreur lors de l'enregistrement : {e}")

    result.salaries_traites = len(salaries_traites)
    return result
